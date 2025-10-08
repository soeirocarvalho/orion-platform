#!/usr/bin/env python3
"""
Quick analysis of Signals sheet structure using openpyxl
"""

import pandas as pd
import os
from openpyxl import load_workbook

def quick_structure_analysis():
    """Quick analysis of Signals sheet structure"""
    
    excel_file = 'attached_assets/ORION_full_curated_megatrends_trends_20250927_195602_1759009168443.xlsx'
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return
    
    try:
        # Use openpyxl for faster header reading
        wb = load_workbook(excel_file, read_only=True)
        print(f'📋 Available sheets: {wb.sheetnames}')
        
        if 'Signals' in wb.sheetnames:
            ws = wb['Signals']
            
            # Get dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            print(f'📊 Signals sheet dimensions: {max_row} rows x {max_col} columns')
            
            # Read header row
            headers = []
            for col in range(1, max_col + 1):
                cell_value = ws.cell(1, col).value
                if cell_value:
                    headers.append(str(cell_value))
                else:
                    headers.append(f"Column_{col}")
            
            print(f'📋 Column headers ({len(headers)}):')
            for i, header in enumerate(headers):
                print(f'  {i+1:2d}. {header}')
            
            # Read first few data rows for samples
            print(f'\n📄 Sample data (first 3 rows):')
            for row_num in range(2, min(5, max_row + 1)):
                row_data = []
                for col in range(1, min(6, max_col + 1)):  # Just first 5 columns
                    cell_value = ws.cell(row_num, col).value
                    row_data.append(str(cell_value)[:30] if cell_value else '')
                print(f'  Row {row_num-1}: {row_data}')
            
            wb.close()
            print(f'\n✅ Quick analysis complete!')
            
        else:
            print('❌ No Signals sheet found')
            wb.close()
        
    except Exception as e:
        print(f'❌ Error: {e}')

if __name__ == "__main__":
    quick_structure_analysis()